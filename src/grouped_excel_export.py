# =============================================================================
# SOURCE MIRROR - READ ONLY REFERENCE COPY
# -----------------------------------------------------------------------------
# Extracted from: Geoprocessing Tool Grouping Excel Export.atbx
#   archive path: GeoprocessingToolGroupingExcelExportV14.tool/tool.script.execute.py
#
# Published so the code is readable and searchable on GitHub. The .atbx is the
# runnable artifact -- editing this file does NOT change the tool. Edit the
# script inside the toolbox in ArcGIS Pro, then re-run:
#     python tools/extract_atbx_scripts.py
# =============================================================================

"""
ArcGIS Pro Geoprocessing Tool: Grouped Excel Export

This script is designed to be used as a Script Tool in an ArcGIS Pro toolbox (.tbx).
It exports attribute data from an input feature layer/table to an Excel workbook (.xlsx),
optionally:

- exporting only a user-selected subset of fields (or all fields)
- grouping rows by a field into separate worksheets (or exporting to a single sheet)
- converting coded domain values to their descriptions (from a geodatabase or a service)
- using field aliases as Excel column headers
- adding a Table of Contents worksheet with hyperlinks (multi-sheet mode)
- auto-fitting column widths and freezing the header row

Notes for publishers/maintainers
-------------------------------
- The tool reads inputs via arcpy.GetParameter* in a fixed index order. If you change the
  toolbox parameter order, you MUST update get_user_parameters_from_tool().
- In "sheets" mode, Excel worksheet names are sanitized and de-duplicated to comply with
  Excel limits (illegal characters + 31 character maximum).
- Domain description mapping supports both:
    * enterprise/file geodatabase domains (arcpy.da.ListDomains)
    * feature services (via ArcGIS API for Python, if available in the Pro environment)

Author: (Ryan Hensel / City of Aspen Utilities Department)
"""
# -----------------------------------------------------------------------------
# Geoprocessing messaging helpers
# -----------------------------------------------------------------------------
# These wrappers standardize messages sent to the ArcGIS Pro Geoprocessing pane.
# NOTE: arcpy is imported below; this is safe because these functions are not
# executed until they are called at runtime.

def info(msg):
    """Send an informational message to the ArcGIS Geoprocessing window."""
    arcpy.AddMessage(f"INFO | {msg}")

def warn(msg):
    """Send a warning message (yellow) to the Geoprocessing window."""
    arcpy.AddWarning(f"WARN | {msg}")

def err(msg):
    """Send an error message (red) to the Geoprocessing window."""
    arcpy.AddError(f"ERROR | {msg}")

def debug(msg):
    """Send a debug message if DEBUG logging is enabled."""
    if DEBUG:
        arcpy.AddMessage(f"DEBUG | {msg}")

def section(title, level="info"):
    """Emit a section header."""
    label = f"[{title}]"
    if level == "warn":
        warn(label)
    elif level == "error":
        err(label)
    else:
        info(label)

def detail(msg, level="info"):
    """Emit an indented detail line."""
    line = f"  {msg}"
    if level == "warn":
        warn(line)
    elif level == "error":
        err(line)
    else:
        info(line)

def kv(key, value, level="info"):
    """Emit a key/value detail line."""
    detail(f"{key}: {value}", level=level)

def emit_table(headers, rows, level="info"):
    """Emit a simple text table with aligned columns."""
    if not rows:
        detail("(none)", level=level)
        return

    widths = [len(str(h)) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(str(cell)))

    header_line = "  " + "  ".join(str(h).ljust(widths[i]) for i, h in enumerate(headers))
    sep_line = "  " + "  ".join("-" * widths[i] for i in range(len(headers)))

    detail(header_line, level=level)
    detail(sep_line, level=level)
    for row in rows:
        line = "  " + "  ".join(str(row[i]).ljust(widths[i]) for i in range(len(headers)))
        detail(line, level=level)


# -----------------------------------------------------------------------------
# Imports and environment checks
# -----------------------------------------------------------------------------

import arcpy
import pandas as pd
import os
from openpyxl.utils import get_column_letter
import importlib

# Confirm required third-party libraries are present in the active Pro Python environment.
# This fails fast with a clear error message if the user runs the tool in a custom env.
for lib in ["pandas", "openpyxl"]:
    if importlib.util.find_spec(lib) is None:
        msg =(
            f"The Python package '{lib}' is not available in this environment. "
            "Please use the default ArcGIS Pro Python environment or contact GIS admin."
        )
        err(msg)
        raise RuntimeError(msg)

arcpy.env.overwriteOutput = True

section("Start")
detail("Imports ready.")

# -----------------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------------
# Excel worksheet names are limited to 31 characters.
MAX_SHEETNAME = 31
DEBUG = True

# -----------------------------------------------------------------------------
# Parameter handling (Script Tool UI -> Python)
# -----------------------------------------------------------------------------

def _read_layer_fieldinfo(feature_layer):
    """
    Read layer fieldInfo in display order.

    Returns a tuple (fields, source):
      ( [{"name": ..., "alias": ..., "visible": ...}, ...], "cim"/"describe"/"" )
    """
    def _field_rows_from_cim_fields(cim_fields, name_attr="name"):
        fields = []
        for f in cim_fields:
            name = getattr(f, name_attr, None) or getattr(f, "fieldName", None)
            if not name:
                continue
            alias = getattr(f, "alias", None) or getattr(f, "aliasName", None) or name
            visible = bool(getattr(f, "visible", True))
            fields.append({
                "name": name,
                "alias": alias,
                "visible": visible
            })
        return fields

    def _from_cim_field_descriptions(cim_obj, version):
        """
        Field descriptions in CIM capture the Fields view order/alias/visibility.
        They typically live on cim.featureTable.fieldDescriptions (feature layers)
        or cim.table.fieldDescriptions (standalone tables).
        """
        candidates = []
        for attr_name in ("featureTable", "table"):
            candidate = getattr(cim_obj, attr_name, None)
            if candidate is not None:
                candidates.append(candidate)

        candidates.append(cim_obj)

        for candidate in candidates:
            cim_desc = getattr(candidate, "fieldDescriptions", None) or []
            if not cim_desc:
                continue

            debug(f"[Fields] cim {version} fieldDescriptions: {len(cim_desc)}")
            fields = _field_rows_from_cim_fields(cim_desc, name_attr="fieldName")
            if fields:
                debug(f"[Fields] source: cim {version} fieldDescriptions")
                return fields, "cim"
        return [], ""

    def _from_cim(layer_obj):
        for version in ("V3", "V2"):
            try:
                cim = layer_obj.getDefinition(version)
            except Exception:
                continue

            cim_fields = getattr(cim, "fieldInfos", None) or []
            if cim_fields:
                fields = _field_rows_from_cim_fields(cim_fields, name_attr="name")
                if fields:
                    debug(f"[Fields] source: cim {version} fieldInfos")
                    return fields, "cim"
            else:
                debug(f"[Fields] cim {version} fieldInfos: 0")

            fields, _ = _from_cim_field_descriptions(cim, version)
            if fields:
                return fields, "cim"

        return [], ""

    def _find_layer_in_project(layer_name, layer_catalog_path):
        try:
            proj = arcpy.mp.ArcGISProject("CURRENT")
        except Exception:
            return None

        for m in proj.listMaps():
            for lyr in m.listLayers():
                if layer_name and (lyr.name == layer_name or lyr.longName == layer_name):
                    return lyr

                if layer_catalog_path:
                    try:
                        lyr_desc = arcpy.Describe(lyr)
                        lyr_catalog = getattr(lyr_desc, "catalogPath", None)
                    except Exception:
                        lyr_catalog = None

                    try:
                        lyr_data_source = lyr.dataSource
                    except Exception:
                        lyr_data_source = None

                    if layer_catalog_path == lyr_catalog or layer_catalog_path == lyr_data_source:
                        return lyr
        return None

    desc = None
    layer_catalog_path = None
    layer_name = None

    try:
        desc = arcpy.Describe(feature_layer)
        layer_catalog_path = getattr(desc, "catalogPath", None)
        layer_name = getattr(desc, "name", None)
    except Exception:
        desc = None

    # Prefer CIM fieldInfos when possible (layer overrides for order/aliases).
    if hasattr(feature_layer, "getDefinition"):
        cim_fields, _ = _from_cim(feature_layer)
        if cim_fields:
            return cim_fields, "cim"

    # If there exists only a layer name/path, try to resolve it in the current project.
    if isinstance(feature_layer, str) or layer_name or layer_catalog_path:
        layer_obj = _find_layer_in_project(feature_layer if isinstance(feature_layer, str) else layer_name, layer_catalog_path)
        if layer_obj is not None:
            debug(f"[Fields] cim layer match: {layer_obj.longName}")
            cim_fields, _ = _from_cim(layer_obj)
            if cim_fields:
                return cim_fields, "cim"
        else:
            debug("[Fields] cim layer match: none")

    if desc is None:
        return [], ""

    field_info = getattr(desc, "fieldInfo", None)
    if not field_info:
        return [], ""

    debug("[Fields] source: describe")

    count = None
    for attr in ("count", "fieldCount", "numFields"):
        if hasattr(field_info, attr):
            try:
                count = int(getattr(field_info, attr))
                break
            except Exception:
                pass

    if count is None:
        try:
            count = int(field_info.count)
        except Exception:
            count = 0

    fields = []

    for i in range(count):
        try:
            name = field_info.getFieldName(i)
        except Exception:
            continue

        try:
            alias = field_info.getFieldAlias(i)
        except Exception:
            alias = name

        try:
            visible = bool(field_info.getFieldVisible(i))
        except Exception:
            visible = True

        fields.append({
            "name": name,
            "alias": alias,
            "visible": visible
        })

    return fields, "describe"


def get_layer_fieldinfo_maps(feature_layer):
    """
    Return layer display order and alias/visibility info derived from fieldInfo.
    Falls back to ListFields when fieldInfo is not available.
    """
    field_info, source = _read_layer_fieldinfo(feature_layer)

    if not field_info:
        warn("[Fields] Layer fieldInfo not available; using dataset field order and aliases.")
        field_info = []
        for f in arcpy.ListFields(feature_layer):
            alias = getattr(f, "aliasName", None) or f.name
            field_info.append({
                "name": f.name,
                "alias": alias,
                "visible": True
            })
        source = "dataset"
    elif source == "describe":
        system_names = {
            "objectid",
            "shape",
            "shape_length",
            "shape_area",
            "globalid",
            "fid",
            "shape_len",
            "shape_area_1",
            "shape_length_1",
            "shape_leng",
            "shape_le_1",
            "shape_ar_1",
        }

        non_system = [
            f for f in field_info
            if f.get("name") and f["name"].lower() not in system_names
        ]
        has_real_alias = any(
            (f.get("alias") or f.get("name")) != f.get("name")
            for f in non_system
        )

        if not has_real_alias:
            desc = None
            try:
                desc = arcpy.Describe(feature_layer)
            except Exception:
                desc = None

            dataset_path = getattr(desc, "catalogPath", None) if desc else None
            if dataset_path:
                dataset_aliases = {}
                for f in arcpy.ListFields(dataset_path):
                    dataset_aliases[f.name] = getattr(f, "aliasName", None) or f.name

                field_info = [
                    {
                        "name": f["name"],
                        "alias": dataset_aliases.get(f["name"], f["name"]),
                        "visible": f.get("visible", True)
                    }
                    for f in field_info
                ]
                debug("[Fields] describe aliases looked like placeholders; rebuilt aliases from dataset.")

    ordered_all = [f["name"] for f in field_info]
    ordered_visible = [f["name"] for f in field_info if f.get("visible", True)]
    alias_map = {f["name"]: (f.get("alias") or f["name"]) for f in field_info}

    # Debug snapshot to verify layer-order + aliases coming from fieldInfo.
    debug("[Fields] ordered_all: " + str(ordered_all))
    debug("[Fields] ordered_visible: " + str(ordered_visible))
    debug("[Fields] alias_map_sample: " + str({k: alias_map[k] for k in list(alias_map)[:10]}))

    return field_info, ordered_all, ordered_visible, alias_map

def get_user_parameters_from_tool():
    """
    Read all ArcGIS Pro tool parameters, normalize them, and return as a single dictionary.
    :return: Dictionary called parameters: all Geoprocessing settings.
            {
                "feature_layer": ...,
                "Export_All_Fields": ...,
                "Fields_To_Export": [...],
                "Group_By_Field": ...,
                "Include_Group_Field_In_Output": ...,
                "Use_Domain_Descriptions": ...,
                "Use_Field_Aliases": ...,
                "Output_XLSX": ...,
                "EXPORT_MODE": ...,
                "ADD_TOC": ...,
                "AUTO_WIDTH": ...,
                "FREEZE_TOP": ...,
            }
    """
    feature_layer = arcpy.GetParameter(0)

    Export_All_Fields = arcpy.GetParameter(1)

    fields_text = arcpy.GetParameterAsText(2)
    requested_raw = fields_text
    missing = []
    hidden = []

    field_info, ordered_all, ordered_visible, alias_map = get_layer_fieldinfo_maps(feature_layer)
    visible_set = set(ordered_visible)
    all_set = set(ordered_all)

    Fields_To_Export = []

    if Export_All_Fields:
        if not ordered_visible:
            msg = "No visible fields found in the layer; nothing to export."
            err(msg)
            raise RuntimeError(msg)

        Fields_To_Export = ordered_visible.copy()


    else:

        if fields_text:

            pieces = fields_text.split(";")

            for raw in pieces:

                clean = raw.strip()

                if clean:
                    Fields_To_Export.append(clean)

            if Fields_To_Export:
                missing = [f for f in Fields_To_Export if f not in all_set]
                hidden = [f for f in Fields_To_Export if (f in all_set and f not in visible_set)]

                if missing:
                    warn(f"[Fields] Fields not on layer (skipped): {missing}")

                if hidden:
                    warn(f"[Fields] Fields hidden on layer (skipped): {hidden}")

                requested_set = {f for f in Fields_To_Export if f in visible_set}
                Fields_To_Export = [f for f in ordered_visible if f in requested_set]


    if len(Fields_To_Export) == 0:

        warn("[Fields] No visible fields were selected; falling back to all visible fields.")

        if not ordered_visible:
            msg = "No visible fields found in the layer; nothing to export."
            err(msg)
            raise RuntimeError(msg)

        Fields_To_Export = ordered_visible.copy()


    Group_By_Field = arcpy.GetParameterAsText(3)

    Include_Group_Field_In_Output = arcpy.GetParameter(4)

    Use_Domain_Descriptions = arcpy.GetParameter(5)  # Setting the domains to be exported as their descriptions

    Use_Field_Aliases = arcpy.GetParameter(6)

    Output_XLSX = arcpy.GetParameterAsText(7)

    EXPORT_MODE = arcpy.GetParameterAsText(8)  # "sheets" (one sheet per group) or "single_sheet"
    ADD_TOC = arcpy.GetParameter(9)  # Boolean if the user wants a Table of Contents sheet added to the beginning
    AUTO_WIDTH = arcpy.GetParameter(10)  # Boolean if the user wants to utilize the auto width function in this
    FREEZE_TOP = arcpy.GetParameter(11)  # Boolean if the user wants to freeze the top row of all Excel sheets.

    parameters = {}

    parameters["feature_layer"] = feature_layer
    parameters["Export_All_Fields"] = Export_All_Fields
    parameters["Fields_To_Export"] = Fields_To_Export
    parameters["Group_By_Field"] = Group_By_Field
    parameters["Include_Group_Field_In_Output"] = Include_Group_Field_In_Output
    parameters["Use_Domain_Descriptions"] = Use_Domain_Descriptions
    parameters["Use_Field_Aliases"] = Use_Field_Aliases
    parameters["Output_XLSX"] = Output_XLSX
    parameters["EXPORT_MODE"] = EXPORT_MODE
    parameters["ADD_TOC"] = ADD_TOC
    parameters["AUTO_WIDTH"] = AUTO_WIDTH
    parameters["FREEZE_TOP"] = FREEZE_TOP

    # The following are derived from the layer and useful for later processing.  They are not user selected. 
    parameters["Layer_FieldInfo"] = field_info
    parameters["Layer_Field_Order"] = ordered_all
    parameters["Layer_Visible_Fields"] = visible_set
    parameters["Layer_All_Fields"] = all_set
    parameters["Layer_Alias_Map"] = alias_map

    try:
        layer_desc = arcpy.Describe(feature_layer)
        layer_label = getattr(layer_desc, "name", None) or str(feature_layer)
    except Exception:
        layer_label = str(feature_layer)

    section("Inputs")
    kv("Layer", layer_label)
    kv("Export all fields", Export_All_Fields)
    kv("Fields to export (raw)", requested_raw if requested_raw else "(none)")
    kv("Group by", Group_By_Field)
    kv("Include group field", Include_Group_Field_In_Output)
    kv("Use domain descriptions", Use_Domain_Descriptions)
    kv("Use field aliases", Use_Field_Aliases)
    kv("Output", Output_XLSX)
    kv("Mode", EXPORT_MODE)
    kv("Add TOC", ADD_TOC)
    kv("AutoFit", AUTO_WIDTH)
    kv("Freeze top row", FREEZE_TOP)

    section("Fields")
    kv("Layer fields (total/visible)", f"{len(ordered_all)}/{len(ordered_visible)}")
    detail(
        f"Export fields: requested={len(Fields_To_Export)}; used={len(Fields_To_Export)}; "
        f"missing={len(missing)}; hidden={len(hidden)}"
    )
    if ordered_visible:
        detail("Field order (visible): " + ", ".join(ordered_visible))
    if Fields_To_Export:
        detail("Export fields (resolved): " + ", ".join(Fields_To_Export))
    if alias_map:
        sample_pairs = list(alias_map.items())[:5]
        sample_text = "; ".join([f"{k}->{v}" for k, v in sample_pairs])
        detail("Alias sample: " + sample_text)

    debug("[Inputs] Parameter parse complete.")

    return parameters

# -----------------------------------------------------------------------------
# Excel worksheet helpers
# -----------------------------------------------------------------------------

def _safe_sheet_name(raw):
    """
    :param raw:
    :return:
    Convert any value to a safe Excel sheet name:
      - handle None/NaN
      - remove illegal characters: : \ / ? * [ ]
      - trim to MAX_SHEETNAME
    !! This helper function will be used in a for loop so it only deals with one sheet name at a time. !!
    """
    if pd.isna(raw):
        base_text = "No Group"
    else:
        base_text = str(raw)

    bad_characters = {":", "\\", "/", "?", "*", "[", "]"}

    cleaned = ""

    for ch in base_text:

        if ch not in bad_characters:
            cleaned += ch   #  += is an augmented assignment operator that re-creates the entire string with the

    cleaned = cleaned.strip()

    if not cleaned:
        cleaned = "Unnamed"

    name = cleaned[:MAX_SHEETNAME]  # will cut off the end of the name at the character limit set in the


    return name

def _ensure_unique_sheet_name(sheet_name, seen):
    """
    :param sheet_name: str
        A proposed sheet name that has already been character cleaned.
    :param seen: a SET that has been recording sheet names already used.
    :return: A unique Excel sheet name based on `sheet_name`.

    - Compare the proposed sheet name to the seen set to find duplicates and add a suffix (" (2)", " (3)", …).
    - Enforces Excel's 31-char limit (via MAX_SHEETNAME previously defined).
    """
    sheet_name = (sheet_name or "").strip()

    if not sheet_name:
        sheet_name = "Sheet"

    sheet_name = sheet_name[:MAX_SHEETNAME]

    lower_sheet_name = sheet_name.lower()

    if lower_sheet_name not in seen:
        seen.add(lower_sheet_name)

        return sheet_name


    suffix_number = 2

    while True:

        suffix = f" ({suffix_number})"

        allowed_sheet_name_length = MAX_SHEETNAME - len(suffix)

        trimmed_sheet_name = sheet_name[:allowed_sheet_name_length]

        candidate = f"{trimmed_sheet_name}{suffix}"

        candidate_lower = candidate.lower()

        if candidate_lower not in seen:
            seen.add(candidate_lower)

            return candidate


        suffix_number += 1


def _autofit_worksheet(ws, df_like):
    """
    :param ws:
    :param df_like:
    :return:
    PURPOSE: Set each excel column width based on some logic.
        They have to be wider than the header and wider than the widest cell value, up to a logical limit.
        Also, don't look at every single row if the number of rows is crazy big.
        Limit the scan to 500 unless the number of rows is less than 500, then scan all row cell sizes.
        Finally, add a buffer of 2 excel unit lengths to make it look nicer.

        inputs: ws: an openpyxml worksheet object, the way to make sheets in Excel
                df_like: a pandas dataframe
    """

    total_rows = len(df_like)

    if total_rows < 500:
        sample_rows = total_rows
    else:
        sample_rows = 500

    col_index = 1

    for col_name in df_like.columns:

        header_text = str(col_name)

        max_len = len(header_text)

        this_col = df_like[col_name]

        sample_values = this_col.head(sample_rows)

        sample_as_text = sample_values.astype(str)

        for cell_text in sample_as_text:

            this_len = len(cell_text)

            if this_len > max_len:
                max_len = this_len

        cushion = 2
        width_with_cushion = max_len + cushion

        max_allowed_width = 60

        if width_with_cushion > max_allowed_width:
            final_width = max_allowed_width
        else:
            final_width = width_with_cushion

        excel_col_letter = get_column_letter(col_index)

        col_dim = ws.column_dimensions[excel_col_letter]

        col_dim.width = final_width

        col_index += 1

def normalize_header(name):
    """
    Converts any column or field name into a consistent 'safe' version.

    Parameters
    ----------
    name : any type (usually string)
        The name to normalize. Could come from ArcGIS field names or user input.

    Returns
    -------
    str
        A cleaned version of the name, formatted consistently.
        The result is truncated to 31 characters to comply with Excel sheet name limits.
    """
    s = str(name)

    s = s.strip()

    s = s.replace("_", "-")

    bad_characters = ['[', ']', '*', '?', ':', '/']

    for bad in bad_characters:

        s = s.replace(bad, "")

        # Excel sheet names have a 31-character limit; column headers do not.
    # We still truncate for consistency and to avoid excessively wide headers.
    return s[:31]



# -----------------------------------------------------------------------------
# Domain mapping (coded value -> description)
# -----------------------------------------------------------------------------

def get_gdb_domain_mappings(feature_layer):
    """
    :param feature_layer:
    :return: A field mapping dictionary called mapping.
    It will try several Describe-based paths (dataElement.path, catalogPath, etc.) to locate the workspace.
    { field: {code: description} } from the geodatabase the layer is in
    Should return an empty dictionary if no useable GDB is available or no coded-value
    domains are on requested fields.
    Later a different defined function will use the fact that {} is either populated or empty in an if statement.
    """

    d = arcpy.Describe(feature_layer)

    workspace = None

    try:

        de = None

        if hasattr(d, "dataElement"):
            de = d.dataElement

        if de:

            de_path = getattr(de, "path", None)
            if de_path:
                workspace = de_path

            de_catalog = getattr(de, "catalogPath", None)

            if de_catalog:
                tmp_desc = arcpy.Describe(de_catalog)
                workspace = getattr(tmp_desc, "path", None)

    except Exception:
        workspace = None

    if not workspace:

        try:

            p = getattr(d, "path", None)

            if p:
                workspace = p

            else:
                cp = getattr(d, "catalogPath", None)

                if cp:
                    tmp = arcpy.Describe(cp)
                    workspace = getattr(tmp, "path", None)

        except Exception:
            workspace = None


    if not workspace:
        return {}

    try:
        all_domains = arcpy.da.ListDomains(workspace)
        if all_domains is None:
            all_domains = []
        if not all_domains:  # [] is falsy; runs when there are zero domains
            debug("[Domains] No domains detected in workspace.")

    except Exception:
        return {}

    domains_by_name = {}

    for domain_obj in all_domains:
        domains_by_name[domain_obj.name] = domain_obj

    fields_objects = arcpy.ListFields(feature_layer) # list of arcpy field objects

    fields_by_name = {}

    for f in fields_objects:
        fields_by_name[f.name] = f  # f.name is the exact field name in the desired layer



    mapping = {}

    for item in fields_by_name.items():
        field_name = item[0]
        field_object = item[1]

        domain_name = getattr(field_object, "domain", None)

        if not domain_name:
            continue # Move the loop to the next field in the fields_by_name dictionary.

        domain_object = domains_by_name.get(domain_name)

        if domain_object is None:
            continue

        if not hasattr(domain_object, "codedValues") or not domain_object.codedValues:
            continue

        code_to_description = {}

        for pair in domain_object.codedValues.items():
            code = pair[0]
            description = pair[1]

            code_to_description[code] = description

        mapping[field_name] = code_to_description

    return mapping

def get_service_domain_mappings_from_layer(feature_layer):
    """
    Read coded value domains from a feature service.

    This requires the ArcGIS API for Python to be available in the ArcGIS Pro
    Python environment, and uses GIS("pro") (the current Pro session) for auth.
    If those prerequisites are not met, this function returns {}.

    :param feature_layer:
    :return:
    Goal: return a nested dictionary like this: { field_name: { code: description } }
    If the layer is NOT a service or if the service has no coded domains: return {}
    Later a different defined function will use the fact that {} is either populated or empty in an if statement.
    """

    DEBUG = True

    d = arcpy.Describe(feature_layer)

    def _first_http(*values): # the * means that there can be any number of variables used in this when called

        for v in values:

            if not isinstance(v, str) or not v:
                continue

            if v.startswith(("http://", "https://")):
                return v

            i = v.find("http://")
            j = v.find("https://")

            positions = []

            if i != -1:
                positions.append(i)

            if j != -1:
                positions.append(j)

            if positions:
                k = min(positions)
            else:
                k = -1

            if k != -1:
                return v[k:]

        return None # end of Step 2

    raw_url = _first_http(
        getattr(d, "dataSource", None),   # often holds a clean service URL
        getattr(d, "catalogPath", None),  # sometimes contains URL in text
        getattr(d, "path", None)          # last way to find the url
    )

    if DEBUG:
        debug(f"[Domains] service-map raw_url candidate: {raw_url}")

    if not raw_url:
        return{}

    url = raw_url.strip()

    def _looks_like_layer(u):

        parts = u.split("/")

        if len(parts) < 2:
            return False  # Too few parts for the url to be a Layer

        second_last = parts[-2]
        last = parts[-1]

        if second_last not in ("FeatureServer", "MapServer"):
            return False

        if not last.isdigit():
            return False

        return True

    if url.strip("/").endswith(("FeatureServer", "MapServer")) and not _looks_like_layer(url):
        if DEBUG:  # If the debug switch is true
            debug("[Domains] service-map root service URL detected, defaulting to sublayer '/0'.")

        url = url.rstrip("/") + "/0"

    if DEBUG:
        debug(f"[Domains] service-map normalized URL used: {url}")


    try:

        try:
            from arcgis.gis import GIS
            from arcgis.features import FeatureLayer
        except Exception as import_ex:
            if DEBUG:
                warn(f"[Domains] service-map ArcGIS API import failed: {import_ex}")
            return {}  # Return empty {} if there is an import problem

        gis = GIS("pro")


        fl = FeatureLayer(url, gis)

        props = getattr(fl, "properties", {})

        fields_meta = []

        if isinstance(props, dict):

            if "fields" in props:
                fields_meta = props["fields"]

            else:
                fields_meta = [] # if there was no 'fields' key.

        else:
            fields_meta = []  # props wasn't even a dictionary.

        mapping = {}
        mapped_fields = []  #  For DEBUG printing only.  It will explain what has been mapped.

        for f in fields_meta:
            dom = f.get("domain")

            if dom and dom.get("type") == "codedValue":


                field_name = f.get("name")
                coded_list = dom.get("codedValues") or [] # make an emtpy list if there are no codedValues.

                code_to_desc = {}

                for cv in coded_list:

                    code = cv.get("code")
                    desc = cv.get("name")

                    code_to_desc[code] = desc

                if field_name:

                    mapping[field_name] = code_to_desc

                    mapped_fields.append(field_name)

            if DEBUG:
                debug(f"[Domains] service-map fields mapped (codedValue): {sorted(mapped_fields)}")

        return mapping

    except Exception as ex:
        if DEBUG:
            warn(f"[Domains] service-map failed to read service domains: {ex}")
        return {} # empty return if anything in step 4 fails.


def get_domains_auto(feature_layer):
    """
    Run the get_gdb_domain_mappings(feature_layer) defined function first and if it returns not empty, then return its
    domains dictionary set to a new variable.
    If it returns empty, then run the get_service_domain_mappings_from_layer(feature_layer) defined function which will
    return either empty or it's domain dictionary set to a new variable.
    """
    gdb_map = get_gdb_domain_mappings(feature_layer)

    if gdb_map:
        return gdb_map, "gdb"

    svc_map = get_service_domain_mappings_from_layer(feature_layer)

    if svc_map:
        return svc_map, "service"

    return {}, None



# -----------------------------------------------------------------------------
# Core pipeline: layer -> DataFrame -> Excel
# -----------------------------------------------------------------------------

def build_dataframe(parameters):
    """
    Build the pandas DataFrame that will be exported to Excel.

    Parameters
    ----------
    parameters : dict
        The dictionary returned by get_user_parameters_from_tool(). At minimum
        must contain:
          - "feature_layer"
          - "Fields_To_Export"
          - "Group_By_Field"
          - "Use_Field_Aliases"

    Returns
    -------
    df_for_export : pandas.DataFrame
        Tabular attribute data only (no geometry), already filtered to the
        requested fields.
    group_field_raw : str or None
        The original name of the group-by field as it appears in the
        feature class, or None if no group-by is used.
    """

    feature_layer = parameters["feature_layer"]
    Fields_To_Export = parameters["Fields_To_Export"]
    Group_By_Field = parameters["Group_By_Field"]
    Include_Group_Field_In_Output = parameters["Include_Group_Field_In_Output"]

    if feature_layer is None:
        msg = "The input feature layer is missing. Please select a layer in the tool."
        err(msg)
        raise RuntimeError(msg)

    if not isinstance(Fields_To_Export, list) or len(Fields_To_Export) == 0:
        msg = "'Fields_To_Export' must be a non-empty Python list of field names."
        err(msg)
        raise RuntimeError(msg)

    if not isinstance(Group_By_Field, str) or not Group_By_Field:
        msg = "Group_By_Field must be a non-empty string"
        err(msg)
        raise RuntimeError(msg)

    layer_field_order = parameters.get("Layer_Field_Order")
    visible_field_set = parameters.get("Layer_Visible_Fields")
    all_field_set = parameters.get("Layer_All_Fields")

    if not layer_field_order or visible_field_set is None or all_field_set is None:
        _, layer_field_order, ordered_visible, _ = get_layer_fieldinfo_maps(feature_layer)
        visible_field_set = set(ordered_visible)
        all_field_set = set(layer_field_order)

    _present = []
    _missing = []
    _hidden = []

    for requested in Fields_To_Export:
        if requested not in all_field_set:
            _missing.append(requested)
        elif requested not in visible_field_set:
            _hidden.append(requested)
        else:
            _present.append(requested)

    if len(_missing) > 0:
        warn(
            f"[Fields] Fields not on layer (skipped): {_missing}"
        )

    if len(_hidden) > 0:
        warn(
            f"[Fields] Fields hidden on layer (skipped): {_hidden}"
        )

    if len(_present) == 0:
        msg = (
            "None of the requested fields are present in the selected layer. "
            "Fix field names or layer selection."
        )
        err(msg)
        raise RuntimeError(msg)

    if Group_By_Field not in all_field_set:
        msg = (
            f"Grouping field '{Group_By_Field}' does not exist in the selected layer. "
            "Check the field name (case-sensitive) or choose a different field."
        )
        err(msg)
        raise RuntimeError(msg)

    group_field_selected = Group_By_Field in _present
    parameters["Group_Field_Selected"] = group_field_selected

    # Debugging messages to find what is happening with layer source issue.
    desc = arcpy.Describe(feature_layer)
    section("Layer")
    kv("dataType", desc.dataType)
    kv("catalogPath", getattr(desc, "catalogPath", None))
    kv("dataSource", getattr(desc, "dataSource", None))
    kv("definitionQuery", getattr(desc, "definitionQuery", None))


    # Detect a layer selection. FIDSet is a semicolon-delimited string of OIDs for selected rows.
    # SearchCursor honors selections on feature layers in Pro.
    fid_set = getattr(desc, "FIDSet", "")
    has_selection = bool(fid_set)

    selection_count = 0
    if has_selection:
        selection_count = len([v for v in fid_set.split(";") if v])
        detail(f"selection: {selection_count} features (exporting selection only)")
    else:
        detail("selection: none (exporting all features)")


    _cursor_fields = _present.copy()

    if Group_By_Field not in _cursor_fields:
        _cursor_fields.append(Group_By_Field)

    include_group_in_output = Include_Group_Field_In_Output or group_field_selected

    if layer_field_order:
        _output_columns = [
            f for f in layer_field_order
            if f in _present or (include_group_in_output and f == Group_By_Field)
        ]
    else:
        _output_columns = _present.copy()
        if include_group_in_output and Group_By_Field not in _output_columns:
            _output_columns.append(Group_By_Field)

    section("Fields")
    detail("Cursor fields: " + ", ".join(_cursor_fields))
    detail("Initial export columns: " + ", ".join(_output_columns))

    rows = []

    with arcpy.da.SearchCursor(feature_layer, _cursor_fields) as cur:
        for rec in cur:
            rows.append(rec)

    df_all = pd.DataFrame(data=rows, columns=_cursor_fields)

    cols_to_keep = []  # New list of Column NAMES to export in Excel.

    for c in _output_columns:
        if c in df_all.columns:
            cols_to_keep.append(c)

    group_only_for_grouping = False
    if Group_By_Field in df_all.columns and Group_By_Field not in cols_to_keep:
        cols_to_keep.append(Group_By_Field)
        group_only_for_grouping = True

    detail("Final export columns: " + ", ".join(cols_to_keep))
    if group_only_for_grouping and (not Include_Group_Field_In_Output) and (not group_field_selected):
        detail("Note: group-by field included for grouping only (not output).")

    df_for_export = df_all[cols_to_keep].copy()

    section("Data")
    kv("Rows exported", len(df_for_export))
    kv("Columns exported", len(df_for_export.columns))

    section("Data Preview")
    preview_cols = list(df_for_export.columns)[:4]
    preview_df = df_for_export.loc[:, preview_cols]
    detail("Preview columns: " + ", ".join(preview_cols))
    preview_text = preview_df.head(5).to_string()
    detail(preview_text)
    remaining_cols = len(df_for_export.columns) - len(preview_cols)
    if remaining_cols > 0:
        detail(f"... {remaining_cols} more column(s) not shown")

    parameters["Group_Field_Visible"] = Group_By_Field in visible_field_set

    group_field_raw = Group_By_Field
    return df_for_export, group_field_raw


def apply_domain_descriptions_if_needed(df, parameters):
    """
    In fields to export that have domain codes applied, convert the codes to their descriptions and format for export.
    :param df:
    :param parameters:
    :return: DataFrame with domain descriptions applied where applicable.
    """

    feature_layer = parameters["feature_layer"]
    Use_Domain_Descriptions = parameters["Use_Domain_Descriptions"]

    if df is None:
        msg = "Input DataFrame 'df' is None. Check that build_dataframe completed successfully."
        err(msg)
        raise RuntimeError(msg)

    if not isinstance(df, pd.DataFrame):
        msg = "Input 'df' must be a pandas DataFrame. Check that build_dataframe returned a DataFrame."
        err(msg)
        raise RuntimeError(msg)

    if not Use_Domain_Descriptions:
        section("Domains")
        detail("Use domain descriptions: False; skipping mapping.")
        return df

    else:
        try:
            domain_map, source = get_domains_auto(feature_layer)
        except Exception as ex:
            domain_map, source = {}, None
            warn(f"[Domains] Could not build domain mappings; exporting coded values as-is. Details: {ex}")

        if not domain_map:
            section("Domains")
            warn("No domain descriptions found (neither geodatabase nor service). Exporting coded values as-is.")

        else:
            section("Domains")
            kv("Source", source)
            kv("Fields with domains", len(domain_map))

            mapped_columns = []
            domain_rows = []

            def _map_with_gentle_fallback(series, code_to_desc):
                """
                Replace coded domain values in a pandas Series with their descriptions.

                WHY THIS IS NEEDED:
                - For TEXT-coded domains, values are usually strings like "A", "B", "C".
                  Mapping works easily.

                - For NUMBER-coded domains, codes might be integers like 1, 2, 3.
                  But when pandas reads a column that has ANY missing values (blank/None),
                  it often converts the whole column to floats:
                      1 becomes 1.0
                      2 becomes 2.0
                  because normal integer columns cannot store NaN.
                  Then:
                      1.0 does NOT match domain key 1
                      "1.0" does NOT match domain key "1"
                  so mapping silently fails.

                This function tries mapping in multiple passes:
                  Pass 1: direct mapping (exact type match)
                  Pass 2: mapping by string (good for many cases)
                  Pass 3: handle the "1.0 should be 1" situation for integer-coded domains
                """

                s1 = series.map(code_to_desc)

                if s1.isna().any():  # .isna() creates a True/False mask; .any() checks if any are True.
                    str_key_map = {str(k): v for k, v in code_to_desc.items()}

                    need = s1.isna()

                    if need.any():
                        s2 = series.astype(str).map(str_key_map)
                        s1 = s1.where(~need, s2)

                # If there are still unmapped values, try a special numeric fix.
                if s1.isna().any():

                    # Integer codes might have been converted to decimal numbers by pandas like 1.0 and 2.0, etc.
                    need = s1.isna()

                    if need.any():

                        # Convert the series to numeric (this will turn "1.0" into 1, but leave non-numeric as NaN).
                        num = pd.to_numeric(series, errors="coerce")

                        # whole means numeric values with no decimal.
                        whole = num.notna() & (num % 1 == 0)

                        if whole.any():

                            # Build a mapping with integer keys.
                            int_key_map = {}
                            for k, v in code_to_desc.items():
                                try:
                                    int_k = int(k)
                                    int_key_map[int_k] = v
                                except Exception:
                                    pass  # If k cannot be converted to int, skip it.

                            # Convert ONLY the whole-number values into a nullable integer series.
                            # Use Int64 which is pandas' nullable integer type.
                            num_as_int = num.where(whole).astype("Int64")

                            # Map the integer codes using the int_key_map.
                            s3 = num_as_int.map(int_key_map)

                            # Fill any still-missing results in s1 with s3.
                            s1 = s1.where(~need, s3)                
        
                # FINAL STEP: Keep original values where no mapping was found.
                s_out = s1.fillna(series)

                return s_out

            for field_name in domain_map.keys():  # Use .keys to just grab the dictionary keys of the domain_map dictionary.

                domain_lookup = domain_map[field_name]

                if field_name not in df.columns:
                    domain_rows.append([field_name, "No", "No (not in export)"])
                    continue  # skip to next field_name

                # Big debugging help print line.
                debug(f"[Domains] field '{field_name}': df dtype={df[field_name].dtype}, sample={df[field_name].head(5).tolist()}")

                col = df[field_name]

                new_col = _map_with_gentle_fallback(col, domain_lookup)

                df[field_name] = new_col

                mapped_columns.append(field_name)
                domain_rows.append([field_name, "Yes", "Yes"])

            if mapped_columns:  # the list that was made just for printing.
                detail("Applied domain descriptions to: " + ", ".join(mapped_columns))
            else:
                detail("No export columns matched a domain field; no mappings applied.")

            emit_table(["Field", "In Export", "Applied"], domain_rows)

        return df

def clean_headers_and_group_column(df, parameters, group_field_raw):
    """
    A basic column header rewrite to replace special characters and underscores.
    Optionally use field aliases as the basis for column headers.
    Also determine the cleaned name of the group-by column so it can be used later.
    :param df: DataFrame returned from build_dataframe (already filtered to export columns)
    :param parameters: dictionary of tool parameters
    :param group_field_raw: original field name used for grouping (as it exists on the layer)
    :return: (df, group_column) where:
             - df has cleaned column headers
             - group_column is the cleaned column name corresponding to group_field_raw
    """

    Use_Field_Aliases = parameters["Use_Field_Aliases"]
    alias_map = parameters.get("Layer_Alias_Map")

    if not alias_map:
        feature_layer = parameters["feature_layer"]
        _, _, _, alias_map = get_layer_fieldinfo_maps(feature_layer)

    if df is None:
        msg = "Input DataFrame 'df' is missing.  Check that build_dataframe ran correctly."
        err(msg)
        raise RuntimeError(msg)

    if not isinstance(df, pd.DataFrame):
        msg = "Input 'df' must be a pandas DataFrame.  Check that build_dataframe ran correctly."
        err(msg)
        raise RuntimeError(msg)

    if df.empty:
        msg = "No rows present for export in the DataFrame after filtering/selection."
        err(msg)
        raise RuntimeError(msg)

    cleaned_columns = []

    if Use_Field_Aliases:

        for col_name in df.columns:

            display_label = alias_map.get(col_name, col_name)

            new_name = normalize_header(display_label)

            cleaned_columns.append(new_name)

        df.columns = cleaned_columns

        group_display_label = alias_map.get(group_field_raw, group_field_raw)

        group_column = normalize_header(group_display_label)

    else:

        for col_name in df.columns:

            new_name = normalize_header(col_name)

            cleaned_columns.append(new_name)

        df.columns = cleaned_columns

        group_column = normalize_header(group_field_raw)

    if group_column not in df.columns:
        msg = (
            f"The group-by column '{group_column}' was not found in the DataFrame after header cleaning.\n"
            f"The columns actually present are: {list(df.columns)}"
        )

        err(msg)
        raise RuntimeError(msg)

    return df, group_column


def export_to_excel(df, group_field, parameters):
    """
    Export the DataFrame to Excel in one of two modes:
      - "single_sheet":  one sheet named "Data"
      - "sheets":        one sheet per distinct group value, plus an optional TOC sheet
    :param df: DataFrame with cleaned headers and final columns for export
    :param group_field: cleaned name of the grouping column (must exist in df.columns)
    :param parameters: dictionary of tool parameters
    :return: None (writes an .xlsx file and registers the output with arcpy)
    """

    Output_XLSX = parameters["Output_XLSX"]
    EXPORT_MODE = parameters["EXPORT_MODE"]
    ADD_TOC = parameters["ADD_TOC"]
    AUTO_WIDTH = parameters["AUTO_WIDTH"]
    FREEZE_TOP = parameters["FREEZE_TOP"]
    Include_Group_Field_In_Output = parameters["Include_Group_Field_In_Output"]
    group_field_selected = parameters.get("Group_Field_Selected", False)
    group_field_visible = parameters.get("Group_Field_Visible", True)

    if not group_field_visible and Include_Group_Field_In_Output:
        warn(
            f"[Fields] Grouping field '{group_field}' is hidden in the layer and will not be exported."
        )
        Include_Group_Field_In_Output = False

    if df is None:
        msg = "Input DataFrame 'df' is missing.  Check that previous steps ran correctly."
        err(msg)
        raise RuntimeError(msg)

    if not isinstance(df, pd.DataFrame):
        msg = "Input 'df' must be a pandas DataFrame.  Check that build_dataframe returned a DataFrame."
        err(msg)
        raise RuntimeError(msg)

    if df.empty:
        msg = "No rows present for export in the DataFrame after filtering/selection."
        err(msg)
        raise RuntimeError(msg)

    if group_field not in df.columns:
        msg = (
            f"The group-by column '{group_field}' was not found in the DataFrame passed to export_to_excel.\n"
            f"The columns actually present are: {list(df.columns)}"
        )
        err(msg)
        raise RuntimeError(msg)

    if not Output_XLSX:
        msg = "Output Excel path is empty. Please specify an .xlsx file in the tool."
        err(msg)
        raise RuntimeError(msg)

    out_dir = os.path.dirname(Output_XLSX)

    if out_dir and not os.path.isdir(out_dir):
        msg = (
            f"Output folder does not exist:\n {out_dir}\n"
            "Please create the folder or choose an existing folder in the tool."
        )
        err(msg)
        raise RuntimeError(msg)

    section("Export")
    kv("Mode", EXPORT_MODE)
    kv("Output", Output_XLSX)
    kv("Group field", group_field)
    kv("Include group field", Include_Group_Field_In_Output)
    kv("Rows", len(df))
    kv("Columns", len(df.columns))
    kv("AutoFit", AUTO_WIDTH)
    kv("Freeze top row", FREEZE_TOP)
    kv("Add TOC", ADD_TOC)

    if EXPORT_MODE == "single_sheet":

        other_cols = []

        for c in df.columns:

            if c != group_field:
                other_cols.append(c)

        sort_keys = [group_field] + other_cols

        df_sorted = df.sort_values(by=sort_keys, kind="stable", ignore_index=True)

        df_to_excel = df_sorted.copy()

        if (not Include_Group_Field_In_Output) and (not group_field_selected) and (group_field in df_to_excel.columns):
            df_to_excel = df_to_excel.drop(columns=[group_field])


        try:
            with pd.ExcelWriter(Output_XLSX, engine="openpyxl") as writer:

                df_to_excel.to_excel(writer, sheet_name="Data", index=False)

                ws = writer.sheets["Data"]

                if AUTO_WIDTH:
                    _autofit_worksheet(ws, df_to_excel)

                if FREEZE_TOP:
                    ws.freeze_panes = ws["A2"]

        except Exception as ex:
            msg = (
                f"Failed writing Excel: {ex}.  You may not have write permission or "
                "an old file is open in Windows"
            )
            err(msg)
            raise RuntimeError(msg)

        arcpy.SetParameter(7, Output_XLSX)

        section("Done")
        detail(f"Export complete (single sheet): {Output_XLSX}")

    elif EXPORT_MODE == "sheets":

        grouped = df.groupby(group_field, dropna=False)  # special GroupBy object!
        kv("Sheets created", grouped.ngroups)

        used_names = set()

        toc_rows = []

        try:

            with pd.ExcelWriter(Output_XLSX, engine="openpyxl") as writer:

                for group_value, g in grouped:

                    base_tab = _safe_sheet_name(group_value)

                    final_sheet_name = _ensure_unique_sheet_name(base_tab, used_names)

                    g_for_sheet = g.copy()

                    if (not Include_Group_Field_In_Output) and (not group_field_selected) and (group_field in g_for_sheet.columns):
                        g_for_sheet = g_for_sheet.drop(columns=[group_field])

                    g_for_sheet.to_excel(writer, sheet_name=final_sheet_name, index=False)

                    ws = writer.sheets[final_sheet_name]

                    if AUTO_WIDTH:
                        _autofit_worksheet(ws, g_for_sheet)

                    if pd.isna(group_value):
                        display_value = "No Group"
                    else:
                        display_value = group_value

                    safe_sheet_ref = final_sheet_name.replace("'", "''")
                    hyperlink_formula = f"=HYPERLINK(\"#'{safe_sheet_ref}'!A1\", \"Go to {final_sheet_name}\")"

                    toc_rows.append({
                        "Group": display_value,
                        "Sheet": final_sheet_name,
                        "Row Count": len(g_for_sheet),
                        "Link": hyperlink_formula
                    })
                    
                if ADD_TOC:

                    toc_df = pd.DataFrame(toc_rows)

                    toc_df = toc_df.sort_values(by=["Group"], kind="stable")

                    toc_df.to_excel(writer, sheet_name="Table of Contents", index=False)

                    workbook = writer.book
                    toc_ws = writer.sheets["Table of Contents"]

                    try:
                        link_col_index = toc_df.columns.get_loc("Link")

                        excel_col = link_col_index + 1

                        for excel_row in range(2, len(toc_df) + 2):
                            cell = toc_ws.cell(row=excel_row, column=excel_col)

                            cell.style = "Hyperlink"

                    except Exception as e:
                        warn(f"Could not apply hyperlink styling: {e}")

                    current_index = workbook.worksheets.index(toc_ws)

                    workbook.move_sheet(toc_ws, offset=-current_index)

                    workbook.active = workbook.worksheets.index(toc_ws)

                    if AUTO_WIDTH:
                        _autofit_worksheet(writer.sheets["Table of Contents"], toc_df)


            arcpy.SetParameter(7, Output_XLSX)

            section("Done")
            detail(f"Export complete (multi-sheet): {Output_XLSX}")

        except Exception as ex:
            msg = f"Failed writing Excel: {ex}.  You may not have write permission or an old file is open in Windows"
            err(msg)
            raise RuntimeError(msg)

    else:
        msg = "Something wrong with export mode selection"
        err(msg)
        raise RuntimeError(msg)


# -----------------------------------------------------------------------------
# Entry point
# -----------------------------------------------------------------------------

def main():
    """
    Orchestrates the whole tool.

    """
    parameters = get_user_parameters_from_tool()
    df, group_field_raw = build_dataframe(parameters)
    df = apply_domain_descriptions_if_needed(df, parameters)
    df, group_field_clean = clean_headers_and_group_column(df, parameters, group_field_raw)
    export_to_excel(df, group_field_clean, parameters)

if __name__ == "__main__":
    main()